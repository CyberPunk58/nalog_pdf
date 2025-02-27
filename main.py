import openpyxl
import datetime
import os
import shutil
import logging
from openpyxl.drawing.image import Image  # Добавлено для работы с изображениями

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Загружаем файл с данными пациентов
patients_file = 'patients.xlsx'  # Укажите путь к вашему файлу
blank_file = 'blank.xlsx'  # Шаблон

# Проверка существования файлов
if not os.path.exists(patients_file):
    raise FileNotFoundError(f"Файл {patients_file} не найден.")
if not os.path.exists(blank_file):
    raise FileNotFoundError(f"Файл {blank_file} не найден.")

# Создаем папку Files, если она не существует
output_folder = 'Files'
if not os.path.exists(output_folder):
    os.makedirs(output_folder)
    logging.info(f"Папка {output_folder} создана.")

try:
    patients_wb = openpyxl.load_workbook(patients_file)
except Exception as e:
    logging.error(f"Ошибка при загрузке файла {patients_file}: {e}")
    exit()

if 'Лист1' not in patients_wb.sheetnames:
    raise ValueError("Лист 'Лист1' не найден в файле patients.xlsx.")

patients_ws = patients_wb['Лист1']  # Получаем конкретный лист с данными (Лист1)

# Функция для определения адреса ячейки по индексу буквы
def get_cell_address(base_col, row, index):
    col_index = openpyxl.utils.column_index_from_string(base_col) + index
    col_letter = openpyxl.utils.get_column_letter(col_index)
    return f"{col_letter}{row}"

# Функция для записи значения в ячейку, учитывая объединённые ячейки
def write_to_cell_safe(ws, cell_address, value):
    cell = ws[cell_address]
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        for merge_range in ws.merged_cells.ranges:
            if cell_address in merge_range:
                top_left_cell = merge_range.min_row, merge_range.min_col
                ws.cell(row=top_left_cell[0], column=top_left_cell[1]).value = value
                return
    else:
        cell.value = value

# Функция для записи текста посимвольно в ячейки
def write_to_cells(base_col, row, text, ws):
    for index, char in enumerate(text):
        cell_address = get_cell_address(base_col, row, index * 2)
        write_to_cell_safe(ws, cell_address, char)

# Функция для записи номера паспорта !!! ПОКА ЭТА ФУНКЦИЯ НЕ ИСПОЛЬЗУЕТСЯ
def write_passport(ws, passport, start_col, row):
    if not passport:  # Пропускаем, если паспорт отсутствует
        return

    passport_str = str(passport)  # Преобразуем номер паспорта в строку
    start_col = start_col  # Начинаем с ячейки AO33
    row = row

    # Записываем первые 4 символа
    for i in range(4):
        if i < len(passport_str):
            cell_address = get_cell_address(start_col, row, i * 2)  # Смещение через одну ячейку
            write_to_cell_safe(ws, cell_address, passport_str[i])

    # Пропускаем две ячейки (смещение на 2 ячейки после первых 4 символов)
    # Записываем оставшиеся символы
    for i in range(4, len(passport_str)):
        cell_address = get_cell_address(start_col, row, i * 2 + 2)  # Смещение + пропуск двух ячеек
        write_to_cell_safe(ws, cell_address, passport_str[i])

# Функция для записи суммы в строку 40 с учетом смещения
def write_amount(ws, amount):
    if amount is None:
        return  # Пропускаем, если сумма отсутствует

    amount_str = str(amount)  # Преобразуем сумму в строку
    start_col = 'BQ'  # Начинаем с ячейки BQ40
    row = 40

    # Вычисляем начальную ячейку с учетом смещения
    start_col_index = openpyxl.utils.column_index_from_string(start_col)
    start_col_index -= len(amount_str) * 2  # Смещаем влево на количество символов суммы * 2
    start_col_letter = openpyxl.utils.get_column_letter(start_col_index)

    # Записываем сумму посимвольно
    for index, char in enumerate(amount_str):
        cell_address = get_cell_address(start_col_letter, row, (index * 2) + 2)
        write_to_cell_safe(ws, cell_address, char)

# Функция для записи даты выдачи паспорта
def write_issue_date(ws, issue_date):
    if not issue_date:  # Пропускаем, если дата отсутствует
        return

    # Преобразуем дату в строку (если это объект datetime)
    if isinstance(issue_date, datetime.datetime):
        issue_date_str = issue_date.strftime('%d.%m.%Y')
    else:
        issue_date_str = str(issue_date)

    # Разделяем дату на день, месяц и год
    day = issue_date_str[:2]  # Первые два символа (день)
    month = issue_date_str[3:5]  # Вторые два символа (месяц)
    year = issue_date_str[6:]  # Последние четыре символа (год)

    # Записываем день (начиная с O35)
    for index, char in enumerate(day):
        cell_address = get_cell_address('O', 35, index * 2)
        write_to_cell_safe(ws, cell_address, char)

    # Записываем месяц (начиная с Q35)
    for index, char in enumerate(month):
        cell_address = get_cell_address('U', 35, index * 2)
        write_to_cell_safe(ws, cell_address, char)

    # Записываем год (начиная с AA35)
    for index, char in enumerate(year):
        cell_address = get_cell_address('AA', 35, index * 2)
        write_to_cell_safe(ws, cell_address, char)

# Функция для записи даты выдачи ВТОРОГО паспорта
def write_issue_date_2(ws, issue_date2):
    if not issue_date2:  # Пропускаем, если дата отсутствует
        return

    # Преобразуем дату в строку (если это объект datetime)
    if isinstance(issue_date2, datetime.datetime):
        issue_date_str = issue_date2.strftime('%d.%m.%Y')
    else:
        issue_date_str = str(issue_date2)

    # Разделяем дату на день, месяц и год
    day = issue_date_str[:2]  # Первые два символа (день)
    month = issue_date_str[3:5]  # Вторые два символа (месяц)
    year = issue_date_str[6:]  # Последние четыре символа (год)

    # Записываем день
    for index, char in enumerate(day):
        cell_address = get_cell_address('O', 23, index * 2)
        write_to_cell_safe(ws, cell_address, char)

    # Записываем месяц
    for index, char in enumerate(month):
        cell_address = get_cell_address('U', 23, index * 2)
        write_to_cell_safe(ws, cell_address, char)

    # Записываем год
    for index, char in enumerate(year):
        cell_address = get_cell_address('AA', 23, index * 2)
        write_to_cell_safe(ws, cell_address, char)

# Функция для записи даты рождения
def write_birthdate(ws, birthdate):
    if not birthdate:  # Пропускаем, если дата отсутствует
        return

    # Преобразуем дату в строку (если это объект datetime)
    if isinstance(birthdate, datetime.datetime):
        birthdate_str = birthdate.strftime('%d.%m.%Y')
    else:
        birthdate_str = str(birthdate)

    # Разделяем дату на день, месяц и год
    day = birthdate_str[:2]  # Первые два символа (день)
    month = birthdate_str[3:5]  # Вторые два символа (месяц)
    year = birthdate_str[6:]  # Последние четыре символа (год)

    # Записываем день (начиная с AY30)
    for index, char in enumerate(day):
        cell_address = get_cell_address('AY', 30, index * 2)
        write_to_cell_safe(ws, cell_address, char)

    # Записываем месяц (начиная с BE30)
    for index, char in enumerate(month):
        cell_address = get_cell_address('BE', 30, index * 2)
        write_to_cell_safe(ws, cell_address, char)

    # Записываем год (начиная с BK30)
    for index, char in enumerate(year):
        cell_address = get_cell_address('BK', 30, index * 2)
        write_to_cell_safe(ws, cell_address, char)

# Функция для записи даты рождения второго человека
def write_birthdate_2(ws, birthdate2):
    if not birthdate2:  # Пропускаем, если дата отсутствует
        return

    # Преобразуем дату в строку (если это объект datetime)
    if isinstance(birthdate2, datetime.datetime):
        birthdate2_str = birthdate2.strftime('%d.%m.%Y')
    else:
        birthdate2_str = str(birthdate2)

    # Разделяем дату на день, месяц и год
    day = birthdate2_str[:2]  # Первые два символа (день)
    month = birthdate2_str[3:5]  # Вторые два символа (месяц)
    year = birthdate2_str[6:]  # Последние четыре символа (год)

    # Записываем день
    for index, char in enumerate(day):
        cell_address = get_cell_address('AY', 18, index * 2)
        write_to_cell_safe(ws, cell_address, char)

    # Записываем месяц
    for index, char in enumerate(month):
        cell_address = get_cell_address('BE', 18, index * 2)
        write_to_cell_safe(ws, cell_address, char)

    # Записываем год
    for index, char in enumerate(year):
        cell_address = get_cell_address('BK', 18, index * 2)
        write_to_cell_safe(ws, cell_address, char)

# Функция для записи периода
def write_period(base_col, row, period_value, ws):
    period_str = str(period_value)
    for index, char in enumerate(period_str):
        cell_address = get_cell_address(base_col, row, index * 2)
        write_to_cell_safe(ws, cell_address, char)

# Функция для записи сегодняшней даты
def write_today_date(ws):
    # Получаем сегодняшнюю дату
    today = datetime.datetime.now()
    day = today.strftime('%d')  # День (две цифры)
    month = today.strftime('%m')  # Месяц (две цифры)
    year = today.strftime('%Y')  # Год (четыре цифры)

    # Записываем день в ячейку V55
    write_to_cells('V', 55, day, ws)

    # Записываем месяц в ячейку AB55
    write_to_cells('AB', 55, month, ws)

    # Записываем год в ячейку AH55
    write_to_cells('AH', 55, year, ws)

# Подготавливаем список для хранения имён новых файлов
new_files = []

# Проходим по каждой строке в файле пациентов, начиная с 2-й строки (0 — это заголовок)
for row in patients_ws.iter_rows(min_row=2, values_only=True):
    (reference_number,
     surname,
     name,
     patronymic,
     birthdate,
     period,
     amount,
     code,
     inn,
     passport,
     issue_date,
     surname2,
     name2,
     patronymic2,
     birthdate2,
     code2,
     inn2,
     passport2,
     issue_date2,
     uploaded) = row

    # Пропускаем строки, которые уже были обработаны
    if uploaded is not None:
        continue

    # Создаем имя нового файла
    date_created = datetime.datetime.now().strftime('%Y-%m-%d')

    # Проверяем, что name и patronymic не равны None
    name_initial = name[0] if name else ''  # Если name None, используем пустую строку
    patronymic_initial = patronymic[0] if patronymic else ''  # Если patronymic None, используем пустую строку

    new_file_name = f'{surname}{name_initial}{patronymic_initial}_{date_created}.xlsx'
    new_file_path = os.path.join(output_folder, new_file_name)  # Сохраняем в папку Files

    # Копируем шаблон
    shutil.copy(blank_file, new_file_path)

    try:
        new_wb = openpyxl.load_workbook(new_file_path)
        new_ws = new_wb.active

        # Записываем сегодняшнюю дату на первый лист
        write_today_date(new_ws)

        # Заполняем фамилию, имя и отчество
        write_to_cells('I', 24, surname, new_ws)  # Фамилия начиная с I24
        write_to_cells('I', 26, name, new_ws)  # Имя начиная с I26
        write_to_cells('I', 28, patronymic, new_ws)  # Отчество начиная с I28

        # Заполняем период
        write_period('BU', 11, period, new_ws)  # Период начиная с BU11

        # Заполняем сумму
        write_amount(new_ws, amount)  # Сумма начиная с BQ40
        write_to_cells('BU', 40, '00', new_ws)  # Копейки начиная с BU40

        # Заполняем ИНН (если есть)
        if inn:
            write_to_cells('I', 30, str(inn), new_ws)  # ИНН начиная с I30

        # Заполняем номер справки (если есть)
        if reference_number:
            write_to_cells('K', 11, str(reference_number), new_ws)  # Номер справки начиная с K11

        # Заполняем код документа
        if code:
            write_to_cells('O', 33, str(code), new_ws)

        # Заполняем паспорт (если есть)
        if passport is not None:  # Проверяем, что passport не равен None
            write_to_cells('AO', 33, str(passport), new_ws)
        else:
            write_to_cells('AO', 33, '', new_ws)  # Записываем пустую строку, если passport равен None

        # Заполняем дату рождения
        write_birthdate(new_ws, birthdate)

        # Заполняем дату выдачи паспорта (если есть)
        write_issue_date(new_ws, issue_date)

        # Если справка на другого человека (surname2 заполнено), записываем данные на лист "Данные ФЛ"
        if surname2:
            # Проверяем, существует ли лист "Данные ФЛ"
            if 'Данные ФЛ' not in new_wb.sheetnames:
                new_wb.create_sheet('Данные ФЛ')  # Создаем лист, если его нет
            fl_ws = new_wb['Данные ФЛ']

            # Записываем фамилию, имя и отчество на лист "Данные ФЛ"
            write_to_cells('I', 12, surname2, fl_ws)
            write_to_cells('I', 14, name2, fl_ws)
            write_to_cells('I', 16, patronymic2, fl_ws)

            # Заполняем паспорт2 (если есть)
            write_to_cells('AO', 21, str(passport2), fl_ws)

            # Заполняем дату выдачи паспорта (если есть)
            write_issue_date_2(fl_ws, issue_date2)

            # Заполняем ИНН2 (если есть)
            if inn2:
                write_to_cells('I', 18, str(inn2), fl_ws)

            # Заполняем дату рождения
            write_birthdate_2(fl_ws, birthdate2)

            # Заполняем код документа
            if code2:
                write_to_cells('O', 21, str(code2), fl_ws)

        # Добавляем изображение на лист
        img = Image('code.png')  # Укажите путь к изображению
        #new_ws.add_image(img, 'C1')  # Укажите ячейку, куда нужно вставить изображение
        #new_ws.

        # Сохраняем изменения
        new_wb.save(new_file_path)
        new_files.append(new_file_name)
        logging.info(f"Файл {new_file_name} успешно создан в папке {output_folder}.")

    except Exception as e:
        logging.error(f"Ошибка при обработке файла {new_file_name}: {e}")
    finally:
        if 'new_wb' in locals():
            new_wb.close()

# Выводим список созданных файлов
logging.info("Созданные файлы:")
for file in new_files:
    logging.info(file)

patients_wb.close()